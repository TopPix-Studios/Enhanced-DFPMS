from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, get_object_or_404
from .models import BackgroundInformation, Profile, Region, Province, City, Barangay, BackgroundInformationLanguage, Specialization,Account, Country , Certificate, PastExperience, Language, Project, Resume
from events.models import RSVP, Attendance, Event
from django.db.models import Count
from django.utils.dateparse import parse_date
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import HttpResponse
import base64
from io import BytesIO
from reportlab.lib.utils import ImageReader
from django.core.exceptions import ValidationError
from collections import defaultdict
import json
from datetime import datetime, timedelta
import re
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count
from django.utils.dateparse import parse_date
from collections import defaultdict
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Image, Spacer
from django.templatetags.static import static
from django.conf import settings
import pandas as pd
from django.utils.timezone import now
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.chart import BarChart, Reference, LineChart
from django.db.models import Count
from .models import BackgroundInformation
from django.contrib.auth.decorators import user_passes_test
from collections import Counter
from openpyxl.cell.cell import MergedCell
from mlxtend.preprocessing import TransactionEncoder
from mlxtend.frequent_patterns import apriori, association_rules
import pandas as pd
from .models import BackgroundInformation
from .report_helpers import (get_skill_breakdowns, get_specialization_breakdowns, get_language_breakdowns)

# 🎨 Color Variables
PRIMARY_COLOR = '00072D'    # Dark Blue
SECONDARY_COLOR = '051650'  # Deep Blue
ACCENT_COLOR = '0A2472'     # Vibrant Blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
HEADER_FILL = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

import os 

def is_admin(user):
    return user.is_staff



def analyze_skills_data(skills_data):
    total_profiles = sum(item['count'] for item in skills_data)
    insights = []

    if total_profiles == 0:
        insights.append("There are no profiles available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to complete their profiles with skills.")
        insights.append("2. Review and improve the data collection process.")
        return insights

    # Add insight about the total number of profiles
    insights.append(f"There are a total of <strong>{total_profiles}</strong> skill users.")

    most_common_skill = max(skills_data, key=lambda x: x['count'])
    least_common_skill = min(skills_data, key=lambda x: x['count'])

    insights.append(f"The most common skill is {most_common_skill['skills__skill']} with <strong>{most_common_skill['count']}</strong> skill users.")
    insights.append(f"The least common skill is {least_common_skill['skills__skill']} with <strong>{least_common_skill['count']}</strong> skill users.")

    # Sort skills_data by count in descending order
    sorted_skills_data = sorted(skills_data, key=lambda x: x['count'], reverse=True)

    # Get top 5 skills
    top_5_skills = sorted_skills_data[:5]

    # Generate insights for the top 5 skills
    for item in top_5_skills:
        skill = item['skills__skill']
        count = item['count']
        percentage = (count / total_profiles) * 100
        insights.append(f"{skill} represents <strong>{count}</strong> profiles, which is <strong>{percentage:.2f}%</strong> of the total skill users.")

    # --- APRIORI SECTION STARTS HERE ---
    # Build transactions
    transactions = []
    background_qs = BackgroundInformation.objects.prefetch_related('skills').all()
    for info in background_qs:
        skill_list = [skill.skill for skill in info.skills.all() if skill.skill]
        if skill_list:
            transactions.append(skill_list)
    if transactions:
        te = TransactionEncoder()
        te_ary = te.fit(transactions).transform(transactions)
        df = pd.DataFrame(te_ary, columns=te.columns_)

        frequent_itemsets = apriori(df, min_support=0.03, use_colnames=True)
        rules = association_rules(frequent_itemsets, metric="lift", min_threshold=1.0)

        if not rules.empty:
            insights.append("<hr><strong>Event Suggestions Based on Skill Patterns:</strong>")
            for _, row in rules.sort_values(by="lift", ascending=False).head(5).iterrows():
                antecedents = ', '.join(list(row['antecedents']))
                consequents = ', '.join(list(row['consequents']))
                confidence = row['confidence'] * 100
                lift = row['lift']
                suggestion = (
                    f"Users with skills in <strong>{antecedents}</strong> also tend to have "
                    f"<strong>{consequents}</strong> (Confidence: <strong>{confidence:.1f}%</strong>, Lift: <strong>{lift:.2f}</strong>). "
                    "Consider organizing an event that combines these skills."
                )
                insights.append(suggestion)
        else:
            insights.append("No strong skill associations were found for suggesting combined-skill events.")
    # --- END OF APRIORI SECTION ---

    return insights




@user_passes_test(is_admin)
def skills_chart(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    region = request.GET.get('region')
    province = request.GET.get('province')
    city = request.GET.get('city')
    barangay = request.GET.get('barangay')
    skills = request.GET.getlist('skill')

    # Get all BackgroundInformation objects
    background_info_qs = BackgroundInformation.objects.all()

    # Filter by date range if provided
    if start_date:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__gte=parse_date(start_date))
    if end_date:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__lte=parse_date(end_date))

    # Filter by location if provided
    if barangay:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__barangay_id=barangay)
    elif city:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__city_id=city)
    elif province:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__province_id=province)
    elif region:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__region_id=region)

    # Filter by skill if provided
    # Filter by skills if provided
    if skills:
        background_info_qs = background_info_qs.filter(skills__skill__in=skills).distinct()

    # Query to get the count of each skill
    skills_count = (
        background_info_qs
        .values('skills__skill')
        .annotate(count=Count('profile__df_id'))
        .order_by('-count')
    )
        
    skills = [entry['skills__skill'] if entry['skills__skill'] is not None else '' for entry in skills_count]
    counts = [entry['count'] for entry in skills_count]

    # Get all regions, provinces, cities, barangays, and unique skills
    regions = Region.objects.all()
    provinces = Province.objects.all()
    cities = City.objects.all()
    barangays = Barangay.objects.all()
    all_skills = BackgroundInformation.objects.values_list('skills__skill', flat=True).distinct()

    # Generate insights
    insights = analyze_skills_data(list(skills_count))
    breakdowns = get_skill_breakdowns(background_info_qs)

    today = now().date()
    context = {
        'skills': skills,
        'counts': counts,
        'regions': regions,
        'selected_skills': skills,  # for template use
        'provinces': provinces,
        'cities': cities,
        'barangays': barangays,
        'all_skills': all_skills,
        'insights': insights,
        'today':  today,
        'gender_breakdown': breakdowns['gender_breakdown'],
        'barangay_breakdown': breakdowns['barangay_breakdown'],
        'age_breakdown': breakdowns['age_breakdown'],
    }
    
    return render(request, 'admin/skills_chart.html', context)


@staff_member_required
def export_skills_to_pdf(request):
    # Get filter parameters
    start_date = request.POST.get('start_date', 'ALL')
    end_date = request.POST.get('end_date', 'ALL')
    region = request.POST.get('region', 'ALL')
    province = request.POST.get('province', 'ALL')
    city = request.POST.get('city', 'ALL')
    barangay = request.POST.get('barangay', 'ALL')

    # Get all BackgroundInformation objects and apply filters
    background_info_qs = BackgroundInformation.objects.all()
    
    start_date_parsed = parse_date(start_date) if start_date != 'ALL' else None
    end_date_parsed = parse_date(end_date) if end_date != 'ALL' else None

    if start_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__gte=start_date_parsed)
    if end_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__lte=(end_date_parsed))

    if barangay and barangay != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__barangay_id=barangay)
    elif city and city != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__city_id=city)
    elif province and province != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__province_id=province)
    elif region and region != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__region_id=region)

    # Query to get the count of each skill
    skills_count = background_info_qs.values('skills__skill').annotate(count=Count('skills')).order_by('-count')
    # Prepare data for the frontend
    data = {
        'start_date': start_date,
        'end_date': end_date,
        'region': region,
        'province': province,
        'city': city,
        'barangay': barangay,
        'skills_count': list(skills_count),  # Convert queryset to a list
    }

    return JsonResponse(data)

@staff_member_required
def export_skills_to_excel(request):
    start_date = request.POST.get('start_date', 'ALL')
    end_date = request.POST.get('end_date', 'ALL')
    region = request.POST.get('region', 'ALL')
    province = request.POST.get('province', 'ALL')
    city = request.POST.get('city', 'ALL')
    barangay = request.POST.get('barangay', 'ALL')

    background_info_qs = BackgroundInformation.objects.all()
    start_date_parsed = parse_date(start_date) if start_date != 'ALL' else None
    end_date_parsed = parse_date(end_date) if end_date != 'ALL' else None

    if start_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__gte=start_date_parsed)
    if end_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__lte=end_date_parsed)
    if barangay != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__barangay_id=barangay)
    elif city != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__city_id=city)
    elif province != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__province_id=province)
    elif region != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__region_id=region)

    # Exclude entries with no linked profile
    background_info_qs = background_info_qs.exclude(profile__isnull=True)

    skills_count = list(
        background_info_qs
        .values('skills__skill')
        .annotate(count=Count('skills'))
        .order_by('-count')
    )

    # Get breakdowns per skill
    breakdowns = get_skill_breakdowns(background_info_qs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Skills Count"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    def style_header(row):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 5

    # Main sheet
    ws.append(['Skill', 'Count'])
    style_header(ws[1])
    for skill in skills_count:
        skill_name = skill['skills__skill'] or 'Unspecified'
        ws.append([skill_name, skill['count']])
    auto_width(ws)

    # Bar Chart
    chart = BarChart()
    chart.title = "Skills Count Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Skill"
    chart.height = 10
    chart.width = 20
    chart.style = 10

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "E2")

    # Breakdown sheets
    for sheet_title, data_dict in {
        'Gender Breakdown': breakdowns['gender_breakdown'],
        'Barangay Breakdown': breakdowns['barangay_breakdown'],
        'Age Group Breakdown': breakdowns['age_breakdown']
    }.items():
        ws_breakdown = wb.create_sheet(title=sheet_title)
        ws_breakdown.append(['Skill', 'Category', 'Count'])
        style_header(ws_breakdown[1])
        for skill, category_counts in data_dict.items():
            for category, count in category_counts.items():
                ws_breakdown.append([skill, category, count])
        auto_width(ws_breakdown)

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'skills_count_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def analyze_specializations_data(specializations_data):
    total_profiles = sum(item['count'] for item in specializations_data)
    insights = []

    if total_profiles == 0:
        insights.append("There are no profiles available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to complete their profiles with specializations.")
        insights.append("2. Review and improve the data collection process.")
        return insights

    # Add insight about the total number of profiles
    insights.append(f"There are a total of <strong>{total_profiles}</strong> profiles.")

    most_common_specialization = max(specializations_data, key=lambda x: x['count'])
    least_common_specialization = min(specializations_data, key=lambda x: x['count'])

    insights.append(f"The most common specialization is {most_common_specialization['specialization__specialization']} with <strong>{most_common_specialization['count']}</strong> profiles.")
    insights.append(f"The least common specialization is {least_common_specialization['specialization__specialization']} with <strong>{least_common_specialization['count']}</strong> profiles.")

    # Sort specializations_data by count in descending order
    sorted_specializations_data = sorted(specializations_data, key=lambda x: x['count'], reverse=True)

    # Get top 5 specializations
    top_5_specializations = sorted_specializations_data[:5]

    # Generate insights for the top 5 specializations
    for item in top_5_specializations:
        specialization = item['specialization__specialization']
        count = item['count']
        percentage = (count / total_profiles) * 100
        insights.append(f"{specialization} represents <strong>{count}</strong> profiles, which is <strong>{percentage:.2f}%</strong> of the total profiles.")

    return insights



@staff_member_required
def specialization_chart(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    region = request.GET.get('region')
    province = request.GET.get('province')
    city = request.GET.get('city')
    barangay = request.GET.get('barangay')
    specialization = request.GET.get('specialization')

    # Get all BackgroundInformation objects
    background_info_qs = BackgroundInformation.objects.all()
    # Exclude BackgroundInformation records that have no profile linked
    background_info_qs = background_info_qs.exclude(profile__isnull=True)

    # Filter by date range if provided
    if start_date:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__gte=parse_date(start_date))
    if end_date:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__lte=parse_date(end_date))

    # Filter by location if provided
    if barangay:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__barangay_id=barangay)
    elif city:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__city_id=city)
    elif province:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__province_id=province)
    elif region:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__region_id=region)

    # Filter by specialization if provided
    if specialization:
        background_info_qs = background_info_qs.filter(specialization__specialization=specialization)

    # Query to get the count of each specialization
    specialization_count = background_info_qs.values('specialization__specialization').annotate(count=Count('specialization')).order_by('-count')
    
    specializations = [entry['specialization__specialization'] if entry['specialization__specialization'] is not None else '' for entry in specialization_count]
    counts = [entry['count'] for entry in specialization_count]

    # Get all regions, provinces, cities, barangays, and unique specializations
    regions = Region.objects.all()
    provinces = Province.objects.all()
    cities = City.objects.all()
    barangays = Barangay.objects.all()
    all_specializations = Specialization.objects.values_list('specialization', flat=True).distinct()

    # Generate insights
    insights = analyze_specializations_data(list(specialization_count))
    breakdowns = get_specialization_breakdowns(background_info_qs)

    today = now().date()
    context = {
        'specializations': specializations,
        'counts': counts,
        'regions': regions,
        'provinces': provinces,
        'cities': cities,
        'barangays': barangays,
        'all_specializations': all_specializations,
        'insights': insights,
        'today': today,
        'gender_breakdown': breakdowns['gender_breakdown'],
        'barangay_breakdown': breakdowns['barangay_breakdown'],
        'age_breakdown': breakdowns['age_breakdown'],
    }
    
    return render(request, 'admin/specialization_chart.html', context)


@staff_member_required
def export_specializations_to_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    region = request.GET.get('region')
    province = request.GET.get('province')
    city = request.GET.get('city')
    barangay = request.GET.get('barangay')
    specialization = request.GET.get('specialization')

    background_info_qs = BackgroundInformation.objects.all()

    if start_date:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__gte=parse_date(start_date))
    if end_date:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__lte=parse_date(end_date))
    if barangay:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__barangay_id=barangay)
    elif city:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__city_id=city)
    elif province:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__province_id=province)
    elif region:
        background_info_qs = background_info_qs.filter(profile__account_id__profile__region_id=region)
    if specialization:
        background_info_qs = background_info_qs.filter(specialization__specialization=specialization)

    # Exclude entries without profile
    background_info_qs = background_info_qs.exclude(profile__isnull=True)

    specialization_count = list(
        background_info_qs.values('specialization__specialization')
        .annotate(count=Count('specialization'))
        .order_by('-count')
    )

    # Get breakdowns
    breakdowns = get_specialization_breakdowns(background_info_qs)

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Specializations Count"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def style_header(row):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 5

    # Main sheet
    ws.append(['Specialization', 'Count'])
    style_header(ws[1])
    for entry in specialization_count:
        specialization_name = entry['specialization__specialization'] or 'Unspecified'
        ws.append([specialization_name, entry['count']])
    auto_width(ws)

    # Bar Chart
    chart = BarChart()
    chart.title = "Specialization Count Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Specialization"
    chart.height = 10
    chart.width = 20
    chart.style = 10

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "E2")

    # Add breakdown sheets
    for sheet_title, breakdown_data in {
        'Gender Breakdown': breakdowns['gender_breakdown'],
        'Barangay Breakdown': breakdowns['barangay_breakdown'],
        'Age Group Breakdown': breakdowns['age_breakdown']
    }.items():
        ws_breakdown = wb.create_sheet(title=sheet_title)
        ws_breakdown.append(['Specialization', 'Category', 'Count'])
        style_header(ws_breakdown[1])

        for specialization, category_counts in breakdown_data.items():
            for category, count in category_counts.items():
                ws_breakdown.append([specialization, category, count])

        auto_width(ws_breakdown)

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'specialization_count_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@staff_member_required
def export_specializations_to_pdf(request):
    # Get filter parameters with defaults for display
    start_date = request.POST.get('start_date', 'ALL')
    end_date = request.POST.get('end_date', 'ALL')
    region = request.POST.get('region', 'ALL')
    province = request.POST.get('province', 'ALL')
    city = request.POST.get('city', 'ALL')
    barangay = request.POST.get('barangay', 'ALL')
    specialization = request.POST.get('specialization', 'ALL')
    chart_image_data = request.POST.get('chartImage')

    # Decode the chart image
    if chart_image_data is None:
        return HttpResponse("No chart image provided", status=400)
    chart_image_data = chart_image_data.split(',')[1]
    chart_image = BytesIO(base64.b64decode(chart_image_data))

    # Get all BackgroundInformation objects
    background_info_qs = BackgroundInformation.objects.all()
    # Exclude BackgroundInformation records that have no profile linked
    background_info_qs = background_info_qs.exclude(profile__isnull=True)

    # Filter by date range if provided and not 'ALL'
    start_date_parsed = parse_date(start_date) if start_date != 'ALL' else None
    end_date_parsed = parse_date(end_date) if end_date != 'ALL' else None

    if start_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__gte=start_date_parsed)
    if end_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__lte=end_date_parsed)

    # Filter by location if provided and not 'ALL'
    if barangay and barangay != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__barangay_id=barangay)
    elif city and city != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__city_id=city)
    elif province and province != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__province_id=province)
    elif region and region != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__region_id=region)

    # Filter by specialization if provided and not 'ALL'
    if specialization and specialization != 'ALL':
        background_info_qs = background_info_qs.filter(specialization__specialization=specialization)

    # Query to get the count of each specialization
    specialization_count = background_info_qs.values('specialization__specialization').annotate(count=Count('specialization')).order_by('-count')

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="specializations_report.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15))

    # Add logo and title
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("Specializations Report", styles['CenterTitle'])
    subtitle = Paragraph(f"Date Range: {start_date if start_date != 'ALL' else 'All'} to {end_date if end_date != 'ALL' else 'All'}<br/>"
                         f"Region: {region if region != 'ALL' else 'All'} | Province: {province if province != 'ALL' else 'All'} | "
                         f"City: {city if city != 'ALL' else 'All'} | Barangay: {barangay if barangay != 'ALL' else 'All'}<br/>"
                         f"Specialization: {specialization if specialization != 'ALL' else 'All'}", styles['SubTitle'])

    # Create a nested table for title and subtitle
    title_table = Table([[title], [subtitle]], colWidths=[5.5*inch])
    title_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    # Create a table for logo and nested title table
    header_table = Table([[logo, title_table]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Draw chart image
    elements.append(Image(chart_image, width=6*inch, height=3*inch))

    # Table Data
    table_data = [['Specialization', 'Count']] + [[entry['specialization__specialization'] if entry['specialization__specialization'] is not None else 'Unknown', entry['count']] for entry in specialization_count]

    # Create Table
    table = Table(table_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response


def analyze_user_data(sorted_dates, user_counts):
    insights = []

    if len(user_counts) == 0 or sum(user_counts) == 0:
        insights.append("There are no user registrations during the selected period.")
        insights.append("Suggestions to encourage user registrations:")
        insights.append("1. Consider running targeted marketing campaigns.")
        insights.append("2. Offer promotions or incentives for new users.")
        insights.append("3. Enhance your social media presence to attract more users.")
        insights.append("4. Improve the user experience on your platform to retain visitors.")
        return insights

    # Convert dates to pandas datetime
    dates = pd.to_datetime(sorted_dates)
    df = pd.DataFrame({'date': dates, 'count': user_counts})
    df.set_index('date', inplace=True)

    # Overall Trend
    if df['count'].sum() > 0:
        if df['count'].iloc[-1] > df['count'].mean():
            insights.append("The chart shows a general trend of increasing user registrations over the selected period. This indicates growing interest and engagement with the platform.")
        else:
            insights.append("The chart shows a stable or declining trend in user registrations over the selected period.")

    # Peaks and Troughs
    peaks = df[df['count'] > df['count'].mean() + df['count'].std()]
    troughs = df[df['count'] < df['count'].mean() - df['count'].std()]

    if not peaks.empty:
        peak_dates = peaks.index.strftime('%Y-%m-%d').tolist()
        insights.append(f"Noticeable peaks can be seen during certain months, possibly due to marketing campaigns or other promotional activities. Peak dates: {', '.join(peak_dates)}.")
    if not troughs.empty:
        trough_dates = troughs.index.strftime('%Y-%m-%d').tolist()
        insights.append(f"Troughs may indicate periods of lower activity. Trough dates: {', '.join(trough_dates)}.")

    # Seasonal Variations
    if len(df) > 6:
        monthly_avg = df['count'].resample('M').mean()
        if monthly_avg.std() > monthly_avg.mean() * 0.1:
            insights.append("There may be seasonal patterns in user registrations, such as higher numbers during holiday seasons or specific times of the year relevant to the platform's user base.")

    # Placeholder for External Events
    insights.append("External events such as product launches, feature updates, or social media buzz can significantly impact user registrations, as reflected in the chart.")

    # Highlight values
    for i in range(len(insights)):
        insights[i] = insights[i].replace('<strong>', '').replace('</strong>', '')
        insights[i] = re.sub(r'(\d+)', r'<strong>\1</strong>', insights[i])

    return insights


@staff_member_required
def profile_chart_view(request):
    # Fetch filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Fetch user join date data with optional date range filtering
    accounts = Account.objects.all()
    
    if start_date:
        accounts = accounts.filter(join_date__gte=parse_date(start_date))
    if end_date:
        accounts = accounts.filter(join_date__lte=parse_date(end_date))
    
    join_date_count = defaultdict(int)
    
    for account in accounts:
        join_date = account.join_date.date()
        join_date_count[join_date] += 1
    
    # Sort the data by date
    sorted_dates = sorted(join_date_count.keys())
    user_counts = [join_date_count[date] for date in sorted_dates]

    # Analyze data to generate insights
    insights = analyze_user_data(sorted_dates, user_counts)
    
    context = {
        'join_dates': json.dumps([date.strftime('%Y-%m-%d') for date in sorted_dates]),
        'user_counts': json.dumps(user_counts),
        'insights': insights,
    }
    return render(request, 'admin/profile_chart.html', context)


from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from collections import defaultdict
from .models import Account

# 🎨 Color Variables
PRIMARY_COLOR = '00072D'
SECONDARY_COLOR = '051650'
ACCENT_COLOR = '0A2472'

@staff_member_required
def export_profile_chart_to_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    accounts = Account.objects.all()

    if start_date:
        accounts = accounts.filter(join_date__gte=parse_date(start_date))
    if end_date:
        accounts = accounts.filter(join_date__lte=parse_date(end_date))

    join_date_count = defaultdict(int)

    for account in accounts:
        join_date = account.join_date.date()
        join_date_count[join_date] += 1

    sorted_dates = sorted(join_date_count.keys())
    user_counts = [join_date_count[date] for date in sorted_dates]

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "User Join Date Chart"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.append(['Join Date', 'User Count'])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    for date, count in zip(sorted_dates, user_counts):
        ws.append([date.strftime('%Y-%m-%d'), count])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 5

    # Bar chart
    chart = BarChart()
    chart.title = "User Join Dates Distribution"
    chart.y_axis.title = "Users Joined"
    chart.x_axis.title = "Date"
    chart.height = 10
    chart.width = 20
    chart.style = 10  # Predefined OpenPyXL chart style

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E2")

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'user_join_date_chart_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@staff_member_required
def export_profile_chart_to_pdf(request):
    # Get chart image data from the POST request
    chart_image_data = request.POST.get('chartImage')
    
    if not chart_image_data:
        raise ValidationError("No chart image data provided.")

    # Decode the chart image
    chart_image_data = chart_image_data.split(',')[1]
    chart_image = BytesIO(base64.b64decode(chart_image_data))

    # Get user join date data for the PDF and combine same dates
    user_data = (Profile.objects
                 .values('account_id__join_date')
                 .annotate(count=Count('account_id'))
                 .order_by('account_id__join_date'))

    # Combine dates and counts
    combined_data = {}
    for entry in user_data:
        join_date = entry['account_id__join_date']
        count = entry['count']
        if join_date:
            join_date_str = join_date.strftime('%Y-%m-%d')
        else:
            join_date_str = 'Unknown'
        
        if join_date_str in combined_data:
            combined_data[join_date_str] += count
        else:
            combined_data[join_date_str] = count

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="profile_chart_report.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22, textColor=colors.HexColor('#051650')))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15, textColor=colors.HexColor('#051650')))

    # Add logo and title
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("User Join Date Distribution Report", styles['CenterTitle'])

    # Create a table for logo and title
    header_table = Table([[logo, title]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Draw chart image
    elements.append(Image(chart_image, width=6*inch, height=3*inch))

    # Table Data
    table_data = [['Join Date', 'Count']]
    for join_date_str, count in combined_data.items():
        table_data.append([join_date_str, count])

    # Create Table with custom color scheme
    table = Table(table_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a43bf')),  # Border color for header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FAFAFA')),  # Text color for header
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')),  # Background color for table rows
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('333')),  # Text color for table rows
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1a43bf')),  # Border color for grid
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response

def analyze_gender_data(gender_data):
    total_profiles = sum(item['count'] for item in gender_data)
    insights = []

    if total_profiles == 0:
        insights.append("There are no profiles available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to complete their profiles.")
        insights.append("2. Review and improve the data collection process.")
        return insights

    most_common_gender = max(gender_data, key=lambda x: x['count'])
    least_common_gender = min(gender_data, key=lambda x: x['count'])

    insights.append(f"The most common gender is {most_common_gender['gender']} with <strong>{most_common_gender['count']}</strong> profiles.")
    insights.append(f"The least common gender is {least_common_gender['gender']} with <strong>{least_common_gender['count']}</strong> profiles.")

    gender_distribution = {item['gender']: item['count'] for item in gender_data}
    for gender, count in gender_distribution.items():
        percentage = (count / total_profiles) * 100
        insights.append(f"{gender} represents <strong>{percentage:.2f}%</strong> of the total profiles (<strong>{count}</strong> out of <strong>{total_profiles}</strong>).")

    # Highlight numbers in all insights
    for i in range(len(insights)):
        insights[i] = re.sub(r'(\d+(\.\d+)?)', r'<strong>\1</strong>', insights[i])

    return insights

@staff_member_required
def gender_chart_view(request):
    profiles = Profile.objects.all()
    gender_filter = request.GET.get('gender')
    if gender_filter:
        profiles = profiles.filter(gender=gender_filter)

    gender_data = profiles.values('gender').annotate(count=Count('gender'))
    unique_genders = set(
        gender.strip().lower()
        for gender in Profile.objects.values_list('gender', flat=True)
        if gender
    )


    insights = analyze_gender_data(list(gender_data))

    context = {
        'gender_data': list(gender_data),
        'unique_genders': unique_genders,
        'insights': insights,
    }

    return render(request, 'admin/gender_chart.html', context)

@staff_member_required
def export_gender_chart_to_pdf(request):
    # Get chart image data from the POST request
    chart_image_data = request.POST.get('chartImage')
    
    if not chart_image_data:
        raise ValidationError("No chart image data provided.")

    # Decode the chart image
    chart_image_data = chart_image_data.split(',')[1]
    chart_image = BytesIO(base64.b64decode(chart_image_data))

    # Get gender data for the PDF
    gender_data = Profile.objects.values('gender').annotate(count=Count('gender')).order_by('gender')

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="gender_chart_report.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22, textColor=colors.HexColor('#333')))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15, textColor=colors.HexColor('#333')))

    # Add logo and title
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("Gender Distribution Report", styles['CenterTitle'])

    # Create a table for logo and title
    header_table = Table([[logo, title]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Draw chart image
    elements.append(Image(chart_image, width=6*inch, height=3*inch))

    # Table Data
    table_data = [['Gender', 'Count']]
    for entry in gender_data:
        gender = entry['gender'] if entry['gender'] else 'Unknown'
        count = entry['count']
        table_data.append([gender, count])

    # Create Table with custom color scheme
    table = Table(table_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a43bf')),  # Header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FAFAFA')),   # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')), # Row background color
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#333')),      # Row text color
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1a43bf')),     # Grid border color
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response

@staff_member_required
def export_gender_chart_to_excel(request):
    gender_filter = request.GET.get('gender')

    profiles = Profile.objects.all()
    if gender_filter:
        profiles = profiles.filter(gender=gender_filter)


    gender_data = profiles.values('gender').annotate(count=Count('gender'))

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Gender Distribution"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.append(['Gender', 'Count'])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    for entry in gender_data:
        ws.append([entry['gender'] or 'Unspecified', entry['count']])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 5

    # Bar chart
    chart = BarChart()
    chart.title = "Gender Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Gender"
    chart.height = 10
    chart.width = 20
    chart.style = 10

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E2")

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'gender_distribution_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def analyze_region_data(region_data):
    total_profiles = sum(item['count'] for item in region_data)
    insights = []

    if total_profiles == 0:
        insights.append("There are no profiles available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to complete their profiles.")
        insights.append("2. Review and improve the data collection process.")
        return insights

    most_common_region = max(region_data, key=lambda x: x['count'])
    least_common_region = min(region_data, key=lambda x: x['count'])

    insights.append(f"The most common region is {most_common_region['region__region']} with <strong>{most_common_region['count']}</strong> profiles.")
    insights.append(f"The least common region is {least_common_region['region__region']} with <strong>{least_common_region['count']}</strong> profiles.")

    # Sort region_data by count in descending order
    sorted_region_data = sorted(region_data, key=lambda x: x['count'], reverse=True)

    # Get top 5 regions
    top_5_regions = sorted_region_data[:5]

    # Generate insights for the top 5 regions
    for item in top_5_regions:
        region = item['region__region']
        count = item['count']
        percentage = (count / total_profiles) * 100
        insights.append(f"{region} represents <strong>{percentage:.2f}%</strong> of the total profiles (<strong>{count}</strong> out of <strong>{total_profiles}</strong>).")

    # Highlight numbers in all insights
    for i in range(len(insights)):
        insights[i] = re.sub(r'(\d+)', r'<strong>\1</strong>', insights[i])

    return insights


@staff_member_required
def region_chart_view(request):
    profiles = Profile.objects.all()
    region_filter = request.GET.get('region')
    if region_filter:
        profiles = profiles.filter(region__region=region_filter)

    region_data = profiles.values('region__region').annotate(count=Count('region'))
    unique_regions = Region.objects.values_list('region', flat=True).distinct()

    insights = analyze_region_data(list(region_data))

    context = {
        'region_data': list(region_data),
        'unique_regions': unique_regions,
        'insights': insights,
    }

    return render(request, 'admin/region_chart.html', context)

@staff_member_required
def export_region_chart_to_excel(request):
    region_filter = request.GET.get('region')
    profiles = Profile.objects.all()
    if region_filter:
        profiles = profiles.filter(region__region=region_filter)

    region_data = profiles.values('region__region').annotate(count=Count('region'))

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Region Distribution"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.append(['Region', 'Count'])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    for entry in region_data:
        ws.append([entry['region__region'] or 'Unspecified', entry['count']])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 5

    # Bar chart
    chart = BarChart()
    chart.title = "Region Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Region"
    chart.height = 10
    chart.width = 20
    chart.style = 10

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E2")

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'region_distribution_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@staff_member_required
def export_region_chart_to_pdf(request):
    # Get chart image data from the POST request
    chart_image_data = request.POST.get('chartImage')
    
    if not chart_image_data:
        raise ValidationError("No chart image data provided.")

    # Decode the chart image
    chart_image_data = chart_image_data.split(',')[1]
    chart_image = BytesIO(base64.b64decode(chart_image_data))

    # Get region data for the PDF
    region_data = Profile.objects.values('region__region').annotate(count=Count('region')).order_by('region__region')

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="region_chart_report.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22, textColor=colors.HexColor('#333')))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15, textColor=colors.HexColor('#333')))

    # Add logo and title
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("Region Distribution Report", styles['CenterTitle'])

    # Create a table for logo and title
    header_table = Table([[logo, title]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Draw chart image
    elements.append(Image(chart_image, width=6*inch, height=3*inch))

    # Table Data
    table_data = [['Region', 'Count']]
    for entry in region_data:
        region = entry['region__region'] if entry['region__region'] else 'Unknown'
        count = entry['count']
        table_data.append([region, count])

    # Create Table with custom color scheme
    table = Table(table_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a43bf')),  # Header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FAFAFA')),   # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')), # Row background color
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#333')),      # Row text color
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1a43bf')),     # Grid border color
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response

def analyze_province_data(province_data):
    total_profiles = sum(item['count'] for item in province_data)
    insights = []

    if total_profiles == 0:
        insights.append("There are no profiles available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to complete their profiles.")
        insights.append("2. Review and improve the data collection process.")
        return insights

    most_common_province = max(province_data, key=lambda x: x['count'])
    least_common_province = min(province_data, key=lambda x: x['count'])

    insights.append(f"The most common province is {most_common_province['province__province']} with <strong>{most_common_province['count']}</strong> profiles.")
    insights.append(f"The least common province is {least_common_province['province__province']} with <strong>{least_common_province['count']}</strong> profiles.")

    # Sort province_data by count in descending order
    sorted_province_data = sorted(province_data, key=lambda x: x['count'], reverse=True)

    # Get top 5 provinces
    top_5_provinces = sorted_province_data[:5]

    # Generate insights for the top 5 provinces
    for item in top_5_provinces:
        province = item['province__province']
        count = item['count']
        percentage = (count / total_profiles) * 100
        insights.append(f"{province} represents <strong>{percentage:.2f}%</strong> of the total profiles (<strong>{count}</strong> out of <strong>{total_profiles}</strong>).")

    # Highlight numbers in all insights
    for i in range(len(insights)):
        insights[i] = re.sub(r'(\d+(\.\d+)?)', r'<strong>\1</strong>', insights[i])

    return insights



@staff_member_required
def province_chart_view(request):
    profiles = Profile.objects.all()
    province_filter = request.GET.get('province')
    if province_filter:
        profiles = profiles.filter(province__province=province_filter)

    province_data = profiles.values('province__province').annotate(count=Count('province'))
    unique_provinces = Province.objects.values_list('province', flat=True).distinct()

    insights = analyze_province_data(list(province_data))

    context = {
        'province_data': list(province_data),
        'unique_provinces': unique_provinces,
        'insights': insights,
    }

    return render(request, 'admin/province_chart.html', context)

@staff_member_required
def export_province_chart_to_excel(request):
    province_filter = request.GET.get('province')
    profiles = Profile.objects.all()
    if province_filter:
        profiles = profiles.filter(province__province=province_filter)

    province_data = profiles.values('province__province').annotate(count=Count('province'))

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Province Distribution"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.append(['Province', 'Count'])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    for entry in province_data:
        ws.append([entry['province__province'] or 'Unspecified', entry['count']])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 5

    # Bar chart
    chart = BarChart()
    chart.title = "Province Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Province"
    chart.height = 10
    chart.width = 20
    chart.style = 10

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E2")

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'province_distribution_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@staff_member_required
def export_province_chart_to_pdf(request):
    # Get chart image data from the POST request
    chart_image_data = request.POST.get('chartImage')
    
    if not chart_image_data:
        raise ValidationError("No chart image data provided.")

    # Decode the chart image
    chart_image_data = chart_image_data.split(',')[1]
    chart_image = BytesIO(base64.b64decode(chart_image_data))

    # Get province data for the PDF
    province_data = Profile.objects.values('province__province').annotate(count=Count('province')).order_by('province__province')

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="province_chart_report.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15))

    # Add logo and title
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("Province Distribution Report", styles['CenterTitle'])

    # Create a table for logo and title
    header_table = Table([[logo, title]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Draw chart image
    elements.append(Image(chart_image, width=6*inch, height=3*inch))

    # Table Data
    table_data = [['Province', 'Count']]
    for entry in province_data:
        province = entry['province__province'] if entry['province__province'] else 'Unknown'
        count = entry['count']
        table_data.append([province, count])

    # Create Table
    table = Table(table_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a43bf')),  # Header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FAFAFA')),   # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')), # Row background color
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#333')),      # Row text color
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1a43bf')),     # Grid border color
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response

def analyze_city_data(city_data):
    total_profiles = sum(item['count'] for item in city_data)
    insights = []

    if total_profiles == 0:
        insights.append("There are no profiles available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to complete their profiles.")
        insights.append("2. Review and improve the data collection process.")
        return insights

    most_common_city = max(city_data, key=lambda x: x['count'])
    least_common_city = min(city_data, key=lambda x: x['count'])

    insights.append(f"The most common city is {most_common_city['city__city']} with <strong>{most_common_city['count']}</strong> profiles.")
    insights.append(f"The least common city is {least_common_city['city__city']} with <strong>{least_common_city['count']}</strong> profiles.")

    # Sort city_data by count in descending order
    sorted_city_data = sorted(city_data, key=lambda x: x['count'], reverse=True)

    # Get top 5 cities
    top_5_cities = sorted_city_data[:5]

    # Generate insights for the top 5 cities
    for item in top_5_cities:
        city = item['city__city']
        count = item['count']
        percentage = (count / total_profiles) * 100
        insights.append(f"{city} represents <strong>{percentage:.2f}%</strong> of the total profiles (<strong>{count}</strong> out of <strong>{total_profiles}</strong>).")

    # Highlight numbers in all insights
    for i in range(len(insights)):
        insights[i] = re.sub(r'(\d+(\.\d+)?)', r'<strong>\1</strong>', insights[i])

    return insights

@staff_member_required
def city_chart_view(request):
    profiles = Profile.objects.all()
    city_filter = request.GET.get('city')
    if city_filter:
        profiles = profiles.filter(city__city=city_filter)

    city_data = profiles.values('city__city').annotate(count=Count('city'))
    unique_cities = City.objects.values_list('city', flat=True).distinct()

    insights = analyze_city_data(list(city_data))

    context = {
        'city_data': list(city_data),
        'unique_cities': unique_cities,
        'insights': insights,
    }

    return render(request, 'admin/city_chart.html', context)

@staff_member_required
def export_city_chart_to_excel(request):
    city_filter = request.GET.get('city')
    profiles = Profile.objects.all()
    if city_filter:
        profiles = profiles.filter(city__city=city_filter)

    city_data = profiles.values('city__city').annotate(count=Count('city'))

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "City Distribution"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.append(['City', 'Count'])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    for entry in city_data:
        ws.append([entry['city__city'] or 'Unspecified', entry['count']])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 5

    # Bar chart
    chart = BarChart()
    chart.title = "City Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "City"
    chart.height = 10
    chart.width = 20
    chart.style = 10

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E2")

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'city_distribution_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@staff_member_required
def export_city_chart_to_pdf(request):
    # Get chart image data from the POST request
    chart_image_data = request.POST.get('chartImage')
    
    if not chart_image_data:
        raise ValidationError("No chart image data provided.")

    # Decode the chart image
    chart_image_data = chart_image_data.split(',')[1]
    chart_image = BytesIO(base64.b64decode(chart_image_data))

    # Get city data for the PDF
    city_data = Profile.objects.values('city__city').annotate(count=Count('city')).order_by('city__city')

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="city_chart_report.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15))

    # Add logo and title
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("City Distribution Report", styles['CenterTitle'])

    # Create a table for logo and title
    header_table = Table([[logo, title]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Draw chart image
    elements.append(Image(chart_image, width=6*inch, height=3*inch))

    # Table Data
    table_data = [['City', 'Count']]
    for entry in city_data:
        city = entry['city__city'] if entry['city__city'] else 'Unknown'
        count = entry['count']
        table_data.append([city, count])

    # Create Table
    table = Table(table_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a43bf')),  # Header background color
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FAFAFA')),   # Header text color
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')), # Row background color
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#333')),      # Row text color
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1a43bf')),     # Grid border color
        ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response


def analyze_barangay_data(barangay_data):
    total_profiles = sum(item['count'] for item in barangay_data)
    insights = []

    if total_profiles == 0:
        insights.append("There are no profiles available for the selected filter.")
        insights.append("Suggestions to increase data availability:")
        insights.append("1. Encourage users to complete their profiles.")
        insights.append("2. Review and enhance the data collection process.")
        return insights

    most_common_barangay = max(barangay_data, key=lambda x: x['count'])
    least_common_barangay = min(barangay_data, key=lambda x: x['count'])

    insights.append(
        f'The barangay with the highest number of profiles is <strong>{most_common_barangay["barangay__barangay"]}</strong>, with <strong>{most_common_barangay["count"]}</strong> profiles. '
        f'<a href="/barangay/{most_common_barangay["barangay__barangay"]}/" target="_blank">View Breakdown</a>'
    )
    insights.append(
        f'The barangay with the fewest profiles is <strong>{least_common_barangay["barangay__barangay"]}</strong>, with only <strong>{least_common_barangay["count"]}</strong> profiles. '
        f'<a href="/barangay/{least_common_barangay["barangay__barangay"]}/" target="_blank">View Breakdown</a>'
    )

    # Sort barangay_data by count in descending order
    sorted_barangay_data = sorted(barangay_data, key=lambda x: x['count'], reverse=True)

    # Get top 5 barangays
    top_5_barangays = sorted_barangay_data[:5]

    # Generate insights for the top 5 barangays
    for item in top_5_barangays:
        barangay = item['barangay__barangay']
        count = item['count']
        percentage = (count / total_profiles) * 100
        insights.append(
            f'<strong>{barangay}</strong> accounts for <strong>{percentage:.2f}%</strong> of all profiles '
            f'(<strong>{count}</strong> out of <strong>{total_profiles}</strong>). '
            f'<a href="/barangay/{barangay}/" target="_blank">View Breakdown</a>'
        )

    return insights


@staff_member_required
def barangay_chart_view(request):
    profiles = Profile.objects.all()
    barangay_filter = request.GET.get('barangay')
    if barangay_filter:
        profiles = profiles.filter(barangay__barangay=barangay_filter)

    barangay_data = profiles.values('barangay__barangay').annotate(count=Count('barangay'))
    unique_barangays = Barangay.objects.values_list('barangay', flat=True).distinct()

    insights = analyze_barangay_data(list(barangay_data))

    context = {
        'barangay_data': list(barangay_data),
        'unique_barangays': unique_barangays,
        'insights': insights,
    }

    return render(request, 'admin/barangay_chart.html', context)

@staff_member_required
def export_barangay_chart_to_excel(request):
    barangay_filter = request.GET.get('barangay')
    profiles = Profile.objects.all()
    if barangay_filter:
        profiles = profiles.filter(barangay__barangay=barangay_filter)

    barangay_data = profiles.values('barangay__barangay').annotate(count=Count('barangay'))

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Barangay Distribution"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.append(['Barangay', 'Count'])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    for entry in barangay_data:
        ws.append([entry['barangay__barangay'] or 'Unspecified', entry['count']])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 5

    # Bar chart
    chart = BarChart()
    chart.title = "Barangay Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Barangay"
    chart.height = 10
    chart.width = 20
    chart.style = 10

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E2")

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'barangay_distribution_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@staff_member_required
def export_barangay_chart_to_pdf(request):
    # Get chart image data from the POST request
    chart_image_data = request.POST.get('chartImage')
    
    if not chart_image_data:
        raise ValidationError("No chart image data provided.")

    # Decode the chart image
    chart_image_data = chart_image_data.split(',')[1]
    chart_image = BytesIO(base64.b64decode(chart_image_data))

    # Get barangay data for the PDF
    barangay_data = Profile.objects.values('barangay__barangay').annotate(count=Count('barangay')).order_by('barangay__barangay')

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="barangay_chart_report.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15))

    # Add logo and title
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("Barangay Distribution Report", styles['CenterTitle'])

    # Create a table for logo and title
    header_table = Table([[logo, title]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Draw chart image
    elements.append(Image(chart_image, width=6*inch, height=3*inch))

    # Table Data
    table_data = [['Barangay', 'Count']]
    for entry in barangay_data:
        barangay = entry['barangay__barangay'] if entry['barangay__barangay'] else 'Unknown'
        count = entry['count']
        table_data.append([barangay, count])

    # Create Table
    table = Table(table_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a43bf')),  # Header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FAFAFA')),   # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')), # Row background color
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#333')),      # Row text color
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1a43bf')),     # Grid border color
    ]))


    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response

def analyze_language_data(language_data, proficiency_data):
    total_profiles = sum(item['count'] for item in language_data)
    insights = []

    if total_profiles == 0:
        insights.append("There are no profiles available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to complete their profiles.")
        insights.append("2. Review and improve the data collection process.")
        return insights

    # Add insight about the total number of profiles
    insights.append(f"There are a total of <strong>{total_profiles}</strong> language speakers.")

    # Most and least common languages
    most_common_language = max(language_data, key=lambda x: x['count'])
    least_common_language = min(language_data, key=lambda x: x['count'])

    insights.append(f"The most common language is <strong>{most_common_language['language__language']}</strong> with <strong>{most_common_language['count']}</strong> speakers.")
    insights.append(f"The least common language is <strong>{least_common_language['language__language']}</strong> with <strong>{least_common_language['count']}</strong> speakers.")

    # Sort by language count
    sorted_language_data = sorted(language_data, key=lambda x: x['count'], reverse=True)
    top_5_languages = sorted_language_data[:5]

    # Generate insights for the top 5 languages
    for item in top_5_languages:
        language = item['language__language']
        count = item['count']
        percentage = (count / total_profiles) * 100
        insights.append(f"<strong>{language}</strong> represents <strong>{count}</strong> profiles, which is <strong>{percentage:.2f}%</strong> of total speakers.")

    # Analyzing proficiency levels
    total_proficiency = sum(item['count'] for item in proficiency_data)
    if total_proficiency > 0:
        proficiency_levels = {"basic": 0, "conversational": 0, "fluent": 0, "native": 0}

        for item in proficiency_data:
            proficiency_levels[item['proficiency_level']] += item['count']

        most_common_proficiency = max(proficiency_levels, key=proficiency_levels.get)
        least_common_proficiency = min(proficiency_levels, key=proficiency_levels.get)

        insights.append(f"The most common proficiency level is <strong>{most_common_proficiency}</strong> with <strong>{proficiency_levels[most_common_proficiency]}</strong> speakers.")
        insights.append(f"The least common proficiency level is <strong>{least_common_proficiency}</strong> with <strong>{proficiency_levels[least_common_proficiency]}</strong> speakers.")

        # Compare proficiency levels
        basic_count = proficiency_levels["basic"]
        fluent_count = proficiency_levels["fluent"]
        native_count = proficiency_levels["native"]

        insights.append(f"Proficiency breakdown: <strong>{basic_count}</strong> at basic level, <strong>{fluent_count}</strong> fluent speakers, and <strong>{native_count}</strong> native speakers.")

        # Additional proficiency insights
        if basic_count > fluent_count:
            insights.append("There are more learners (basic speakers) than fluent users. Consider adding language training programs.")
        if native_count > fluent_count:
            insights.append("Native speakers dominate over fluent speakers, indicating a strong linguistic base in this language.")

    return insights


@staff_member_required
def language_chart(request):
    # Get filter parameters
    language = request.GET.get('language')
    region = request.GET.get('region')
    province = request.GET.get('province')
    city = request.GET.get('city')
    barangay = request.GET.get('barangay')

    # Get all BackgroundInformation objects
    background_info_qs = BackgroundInformation.objects.all()
    background_info_qs = background_info_qs.exclude(profile__isnull=True)

    # Apply location-based filtering
    if barangay:
        background_info_qs = background_info_qs.filter(profile__barangay__barangay=barangay)
    elif city:
        background_info_qs = background_info_qs.filter(profile__city__city=city)
    elif province:
        background_info_qs = background_info_qs.filter(profile__province__province=province)
    elif region:
        background_info_qs = background_info_qs.filter(profile__region__region=region)

    # Filter by language if provided
    if language:
        background_info_qs = background_info_qs.filter(language__language=language)

    # Query to get the count of languages
    language_count = background_info_qs.values('language__language').annotate(count=Count('language')).order_by('-count')

    # Query to get proficiency level distribution
    proficiency_distribution = (
        BackgroundInformationLanguage.objects.filter(background_information__in=background_info_qs)
        .values('language__language', 'proficiency_level')
        .annotate(count=Count('proficiency_level'))
        .order_by('-count')
    )


    # Process data for visualization
    languages = [entry['language__language'] if entry['language__language'] else '' for entry in language_count]
    counts = [entry['count'] for entry in language_count]

    proficiency_labels = [
        f"{entry['language__language']} ({entry['proficiency_level']})"
        for entry in proficiency_distribution
    ]
    proficiency_counts = [entry['count'] for entry in proficiency_distribution]

    # Get all languages, regions, provinces, cities, and barangays for the filter dropdowns
    all_languages = Language.objects.all()
    regions = Region.objects.all()
    provinces = Province.objects.all()
    cities = City.objects.all()
    barangays = Barangay.objects.all()

    # Generate insights
    insights = analyze_language_data(list(language_count), list(proficiency_distribution))

    today = now().date()
    breakdowns = get_language_breakdowns(background_info_qs)

    context = {
        'languages': languages,
        'counts': counts,
        'all_languages': all_languages,
        'regions': regions,
        'provinces': provinces,
        'cities': cities,
        'barangays': barangays,
        'proficiency_labels': proficiency_labels,
        'proficiency_counts': proficiency_counts,
        'insights': insights,
        'today': today,
        'gender_breakdown': breakdowns['gender_breakdown'],
        'barangay_breakdown': breakdowns['barangay_breakdown'],
        'age_breakdown': breakdowns['age_breakdown'],
    }

    return render(request, 'admin/language_chart.html', context)

@staff_member_required
def export_language_chart_to_excel(request):
    language = request.GET.get('language')
    region = request.GET.get('region')
    province = request.GET.get('province')
    city = request.GET.get('city')
    barangay = request.GET.get('barangay')

    background_info_qs = BackgroundInformation.objects.all()

    # Apply location filters
    if barangay:
        background_info_qs = background_info_qs.filter(profile__barangay__barangay=barangay)
    elif city:
        background_info_qs = background_info_qs.filter(profile__city__city=city)
    elif province:
        background_info_qs = background_info_qs.filter(profile__province__province=province)
    elif region:
        background_info_qs = background_info_qs.filter(profile__region__region=region)

    # Filter by language if provided
    if language:
        background_info_qs = background_info_qs.filter(language__language=language)

    # Query for language count
    language_count = background_info_qs.values('language__language').annotate(count=Count('language')).order_by('-count')

    # Query for proficiency level distribution
    proficiency_distribution = (
        BackgroundInformationLanguage.objects.filter(background_information__in=background_info_qs)
        .values('language__language', 'proficiency_level')
        .annotate(count=Count('proficiency_level'))
        .order_by('-count')
    )

    # Create Workbook
    wb = Workbook()
    
    # ---- Language Distribution Sheet ----
    ws1 = wb.active
    ws1.title = "Language Distribution"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws1.append(['Language', 'Count'])
    for cell in ws1[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    for entry in language_count:
        ws1.append([entry['language__language'] or 'Unspecified', entry['count']])

    # Auto-adjust column widths
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max_length + 5

    # Bar Chart - Language Distribution
    chart1 = BarChart()
    chart1.title = "Language Distribution"
    chart1.y_axis.title = "Count"
    chart1.x_axis.title = "Language"
    chart1.height = 10
    chart1.width = 20
    chart1.style = 10

    data = Reference(ws1, min_col=2, min_row=1, max_row=ws1.max_row)
    categories = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(categories)

    ws1.add_chart(chart1, "E2")

    # ---- Proficiency Level Sheet ----
    ws2 = wb.create_sheet(title="Proficiency Levels")

    # Header
    ws2.append(['Language', 'Proficiency Level', 'Count'])
    for cell in ws2[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    for entry in proficiency_distribution:
        ws2.append([
            entry['language__language'] or 'Unspecified',
            entry['proficiency_level'].capitalize(),
            entry['count']
        ])

    # Auto-adjust column widths
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max_length + 5

    # Bar Chart - Proficiency Levels
    chart2 = BarChart()
    chart2.title = "Proficiency Level Distribution"
    chart2.y_axis.title = "Count"
    chart2.x_axis.title = "Proficiency Level"
    chart2.height = 10
    chart2.width = 20
    chart2.style = 11

    data = Reference(ws2, min_col=3, min_row=1, max_row=ws2.max_row)
    categories = Reference(ws2, min_col=2, min_row=2, max_row=ws2.max_row)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(categories)

    ws2.add_chart(chart2, "E2")

    # ---- Insights Sheet ----
    ws3 = wb.create_sheet(title="Insights")
    ws3.append(["Generated Insights"])
    
    # Generate insights based on the data
    insights = analyze_language_data(list(language_count), list(proficiency_distribution))
    for insight in insights:
        ws3.append([insight])

    # Auto-adjust column width
    ws3.column_dimensions['A'].width = 80

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'language_distribution_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response



@staff_member_required
def export_languages_to_pdf(request):
    # Get filter parameters with defaults for display
    start_date = request.POST.get('start_date', 'ALL')
    end_date = request.POST.get('end_date', 'ALL')
    region = request.POST.get('region', 'ALL')
    province = request.POST.get('province', 'ALL')
    city = request.POST.get('city', 'ALL')
    barangay = request.POST.get('barangay', 'ALL')
    chart_image_data = request.POST.get('chartImage')

    # Decode the chart image
    if chart_image_data is None:
        return HttpResponse("No chart image provided", status=400)
    chart_image_data = chart_image_data.split(',')[1]
    chart_image = BytesIO(base64.b64decode(chart_image_data))

    # Get all BackgroundInformation objects
    background_info_qs = BackgroundInformation.objects.all()

    # Filter by date range if provided and not 'ALL'
    start_date_parsed = parse_date(start_date) if start_date != 'ALL' else None
    end_date_parsed = parse_date(end_date) if end_date != 'ALL' else now().date()

    if start_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__gte=start_date_parsed)
    if end_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__lte=end_date_parsed)

    # Filter by location if provided and not 'ALL'
    if barangay and barangay != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__barangay_id=barangay)
    elif city and city != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__city_id=city)
    elif province and province != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__province_id=province)
    elif region and region != 'ALL':
        background_info_qs = background_info_qs.filter(profile__account_id__profile__region_id=region)

    # Query to get the count of each language
    language_count = background_info_qs.values('language__language').annotate(count=Count('language')).order_by('-count')

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="languages_report.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15))

    # Add logo and title
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("Languages Report", styles['CenterTitle'])
    subtitle = Paragraph(f"Date Range: {start_date if start_date != 'ALL' else 'All'} to {end_date if end_date != 'ALL' else 'All'}<br/>"
                         f"Region: {region if region != 'ALL' else 'All'} | Province: {province if province != 'ALL' else 'All'} | "
                         f"City: {city if city != 'ALL' else 'All'} | Barangay: {barangay if barangay != 'ALL' else 'All'}", styles['SubTitle'])

    # Create a nested table for title and subtitle
    title_table = Table([[title], [subtitle]], colWidths=[5.5*inch])
    title_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    # Create a table for logo and nested title table
    header_table = Table([[logo, title_table]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Draw chart image
    elements.append(Image(chart_image, width=6*inch, height=3*inch))

    # Table Data
    table_data = [['Language', 'Count']]
    for entry in language_count:
        language = entry['language__language'] if entry['language__language'] is not None else 'Unknown'
        count = entry['count']
        table_data.append([language, count])

    # Create Table
    table = Table(table_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response


def analyze_past_experience_data(country_data):
    total_experiences = sum(item['count'] for item in country_data)
    insights = []

    if total_experiences == 0:
        insights.append("There are no past experiences available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to complete their profiles with past experiences.")
        insights.append("2. Review and improve the data collection process.")
        return insights

    # Add insight about the total number of past experiences
    insights.append(f"There are a total of <strong>{total_experiences}</strong> past experiences.")

    most_common_country = max(country_data, key=lambda x: x['count'])
    least_common_country = min(country_data, key=lambda x: x['count'])

    insights.append(f"The most common country for past experiences is {most_common_country['past_experiences__country__country']} with <strong>{most_common_country['count']}</strong> past experiences.")
    insights.append(f"The least common country for past experiences is {least_common_country['past_experiences__country__country']} with <strong>{least_common_country['count']}</strong> past experiences.")

    # Sort country_data by count in descending order
    sorted_country_data = sorted(country_data, key=lambda x: x['count'], reverse=True)

    # Get top 5 countries
    top_5_countries = sorted_country_data[:5]

    # Generate insights for the top 5 countries
    for item in top_5_countries:
        country = item['past_experiences__country__country']
        count = item['count']
        percentage = (count / total_experiences) * 100
        insights.append(f"{country} represents <strong>{count}</strong> past experiences, which is <strong>{percentage:.2f}%</strong> of the total past experiences.")

    return insights

@staff_member_required
def past_experience_chart(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    country = request.GET.get('country')

    # Get all BackgroundInformation objects
    background_info_qs = BackgroundInformation.objects.all()

    # Filter by date range if provided
    if start_date:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__gte=parse_date(start_date))
    if end_date:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__lte=parse_date(end_date))

    # Filter by country if provided
    if country:
        background_info_qs = background_info_qs.filter(past_experiences__country__country=country)

    # Query to get the count of past experiences by country
    country_count = background_info_qs.values('past_experiences__country__country').annotate(count=Count('past_experiences')).order_by('-count')
    
    countries = [entry['past_experiences__country__country'] if entry['past_experiences__country__country'] is not None else '' for entry in country_count]
    counts = [entry['count'] for entry in country_count]

    # Get all countries for the filter dropdown
    all_countries = Country.objects.all()

    # Generate insights
    insights = analyze_past_experience_data(list(country_count))

    today = now().date()
    context = {
        'countries': countries,
        'counts': counts,
        'all_countries': all_countries,
        'insights': insights,
        'today': today,
    }
    
    return render(request, 'admin/past_experience_chart.html', context)


@staff_member_required
def export_past_experiences_to_pdf(request):
    # Get filter parameters with defaults for display
    start_date = request.POST.get('start_date', 'ALL')
    end_date = request.POST.get('end_date', 'ALL')
    chart_image_data = request.POST.get('chartImage')

    # Decode the chart image
    if chart_image_data is None:
        return HttpResponse("No chart image provided", status=400)
    chart_image_data = chart_image_data.split(',')[1]
    chart_image = BytesIO(base64.b64decode(chart_image_data))

    # Get all BackgroundInformation objects
    background_info_qs = BackgroundInformation.objects.all()

    # Filter by date range if provided and not 'ALL'
    start_date_parsed = parse_date(start_date) if start_date != 'ALL' else None
    end_date_parsed = parse_date(end_date) if end_date != 'ALL' else None

    if start_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__gte=start_date_parsed)
    if end_date_parsed:
        background_info_qs = background_info_qs.filter(profile__account_id__join_date__lte=end_date_parsed)

    # Query to get the count of past experiences by country
    country_count = PastExperience.objects.values('country__country').annotate(count=Count('country')).order_by('-count')

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="past_experiences_report.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15))

    # Add logo and title
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("Past Experiences Report", styles['CenterTitle'])
    subtitle = Paragraph(f"Date Range: {start_date if start_date != 'ALL' else 'All'} to {end_date if end_date != 'ALL' else 'All'}", styles['SubTitle'])

    # Create a nested table for title and subtitle
    title_table = Table([[title], [subtitle]], colWidths=[5.5*inch])
    title_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    # Create a table for logo and nested title table
    header_table = Table([[logo, title_table]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Draw chart image
    elements.append(Image(chart_image, width=6*inch, height=3*inch))

    # Table Data
    table_data = [['Country', 'Count']]
    for entry in country_count:
        country = entry['country__country'] if entry['country__country'] is not None else 'Unknown'
        count = entry['count']
        table_data.append([country, count])

    # Create Table
    table = Table(table_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response


@staff_member_required
def admin_profile_detail(request, profile_id):
    profile = get_object_or_404(Profile, pk=profile_id)
    projects = Project.objects.filter(profile=profile)
    resume = Resume.objects.filter(profile=profile).first()
    certificates = Certificate.objects.filter(profile=profile)
    context = {
        'profile': profile,
        'projects': projects,
        'resume': resume,
        'certificates': certificates,
    }

    return render(request, 'admin/admin_profile_detail.html', context)

def analyze_event_data(dates, counts):
    total_events = sum(counts)
    insights = []

    if total_events == 0:
        insights.append("There are no event attendance records available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to attend more events.")
        insights.append("2. Review and improve the event engagement process.")
        return insights

    insights.append(f"There are a total of <strong>{total_events}</strong> event attendances.")
    
    most_common_date = dates[counts.index(max(counts))]
    least_common_date = dates[counts.index(min(counts))]

    insights.append(f"The date with the highest attendance is <strong>{most_common_date}</strong> with <strong>{max(counts)}</strong> attendances.")
    insights.append(f"The date with the lowest attendance is <strong>{least_common_date}</strong> with <strong>{min(counts)}</strong> attendances.")
    
    return insights

def analyze_age_data(age_ranges, counts):
    total_ages = sum(counts)
    insights = []

    if total_ages == 0:
        insights.append("There are no age distribution records available for the selected filter.")
        insights.append("Suggestions to increase data:")
        insights.append("1. Encourage users to provide their age information.")
        insights.append("2. Review and improve the data collection process.")
        return insights

    age_ranges_dict = {
        'A': '20 below',
        'B': '20-29',
        'C': '30-39',
        'D': '40-49',
        'E': '50+'
    }

    insights.append(f"There are a total of <strong>{total_ages}</strong> recorded ages.")

    most_common_age = age_ranges_dict[age_ranges[counts.index(max(counts))]]
    least_common_age = age_ranges_dict[age_ranges[counts.index(min(counts))]]

    insights.append(f"The most common age range is <strong>{most_common_age}</strong> with <strong>{max(counts)}</strong> occurrences.")
    insights.append(f"The least common age range is <strong>{least_common_age}</strong> with <strong>{min(counts)}</strong> occurrences.")
    
    return insights

@staff_member_required
def event_chart_view(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Filter events based on the date range, if provided
    events = Event.objects.filter(is_published=True)
    if start_date:
        events = events.filter(start_datetime__gte=parse_date(start_date))
    if end_date:
        events = events.filter(end_datetime__lte=parse_date(end_date))

    # Initialize dictionaries to aggregate data
    event_attendance_data = defaultdict(list)
    age_range_data = defaultdict(int)
    total_attendance_count = 0

    # Collect data for each event in the filtered set
    for event in events:
        attendances = Attendance.objects.filter(event=event)
        attendance_count = attendances.count()
        total_attendance_count += attendance_count
        start_date_str = event.start_datetime.date().strftime('%Y-%m-%d')
        
        # Collect attendance data by date
        event_attendance_data[start_date_str].append({
            'title': event.title,
            'count': attendance_count
        })

        # Aggregate age range data across all events
        for attendance in attendances:
            if attendance.age_range:
                age_range_data[attendance.age_range] += 1

    # Prepare lists for chart labels and values
    dates_with_events = sorted(event_attendance_data.keys())
    attendance_counts = [
        sum(item['count'] for item in event_attendance_data[date]) 
        for date in dates_with_events
    ]
    event_titles = [
        ', '.join(item['title'] for item in event_attendance_data[date])
        for date in dates_with_events
    ]

    # Prepare event data list for table display
    event_data_list = [
        (item['title'], date, item['count'])
        for date, events in event_attendance_data.items()
        for item in events
    ]

    # Age range labels and counts for chart display
    age_ranges_dict = {
        'A': '20 below',
        'B': '20-29',
        'C': '30-39',
        'D': '40-49',
        'E': '50+'
    }
    age_ranges = sorted(age_range_data.keys())
    age_range_labels = [age_ranges_dict.get(age_range, age_range) for age_range in age_ranges]
    age_range_counts = [age_range_data[age_range] for age_range in age_ranges]
    age_range_data_list = [(age_ranges_dict.get(age_range, age_range), age_range_data[age_range]) for age_range in age_ranges]

    # Generate insights directly within the view based on calculated data
    event_insights = [
        f"Total attendance across all events: {total_attendance_count}",
        f"Highest single-day attendance: {max(attendance_counts) if attendance_counts else 0}"
    ]
    if age_range_counts:
        max_age_range = age_ranges[age_range_counts.index(max(age_range_counts))]
        event_insights.append(f"Most common age range: {age_ranges_dict.get(max_age_range, max_age_range)}")

    # Context data for the template
    context = {
        'dates': dates_with_events,
        'attendance_counts': attendance_counts,
        'event_titles': event_titles,
        'event_data_list': event_data_list,
        'age_ranges': age_range_labels,
        'age_range_counts': age_range_counts,
        'age_range_data_list': age_range_data_list,
        'event_insights': event_insights,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'admin/event_chart.html', context)

@staff_member_required
def export_event_chart_to_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Filter events based on the date range, if provided
    events = Event.objects.filter(is_published=True)
    if start_date:
        events = events.filter(start_datetime__gte=start_date)
    if end_date:
        events = events.filter(end_datetime__lte=end_date)

    # Initialize data storage
    event_attendance_data = defaultdict(list)
    age_range_data = defaultdict(int)
    total_attendance_count = 0

    # Process events
    for event in events:
        attendances = Attendance.objects.filter(event=event)
        attendance_count = attendances.count()
        total_attendance_count += attendance_count
        start_date_str = event.start_datetime.date().strftime('%Y-%m-%d')

        event_attendance_data[start_date_str].append({
            'title': event.title,
            'count': attendance_count
        })

        # Aggregate age range data
        for attendance in attendances:
            if attendance.age_range:
                age_range_data[attendance.age_range] += 1

    # Prepare lists for chart labels and values
    dates_with_events = sorted(event_attendance_data.keys())
    attendance_counts = [
        sum(item['count'] for item in event_attendance_data[date])
        for date in dates_with_events
    ]

    # Age range data
    age_ranges_dict = {
        'A': '20 below',
        'B': '20-29',
        'C': '30-39',
        'D': '40-49',
        'E': '50+'
    }
    age_ranges = sorted(age_range_data.keys())
    age_range_labels = [age_ranges_dict.get(age_range, age_range) for age_range in age_ranges]
    age_range_counts = [age_range_data[age_range] for age_range in age_ranges]

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Event Attendance Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws.append(["Date", "Event Title", "Attendance Count"])
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Event Data
    for date, events in event_attendance_data.items():
        for item in events:
            ws.append([date, item['title'], item['count']])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 5

    # Bar Chart for Attendance
    chart = BarChart()
    chart.title = "Event Attendance Over Time"
    chart.y_axis.title = "Attendance Count"
    chart.x_axis.title = "Date"
    chart.height = 10
    chart.width = 20
    chart.style = 10

    data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E2")

    # Age Range Sheet
    ws2 = wb.create_sheet(title="Age Range Data")
    ws2.append(["Age Range", "Count"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    for i, age_label in enumerate(age_range_labels):
        ws2.append([age_label, age_range_counts[i]])

    # Age Range Chart
    age_chart = BarChart()
    age_chart.title = "Age Range Distribution"
    age_chart.y_axis.title = "Count"
    age_chart.x_axis.title = "Age Range"
    age_chart.height = 10
    age_chart.width = 20
    age_chart.style = 10

    age_data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
    age_categories = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    age_chart.add_data(age_data, titles_from_data=True)
    age_chart.set_categories(age_categories)

    ws2.add_chart(age_chart, "E2")

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'event_chart_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@staff_member_required
def export_event_age_report_pdf(request):
    start_date = request.POST.get('start_date', 'ALL')
    end_date = request.POST.get('end_date', 'ALL')
    event_chart_image_data = request.POST.get('event_chart_image')
    age_chart_image_data = request.POST.get('age_chart_image')

    if event_chart_image_data is None or age_chart_image_data is None:
        return HttpResponse("No chart images provided", status=400)
    
    event_chart_image = BytesIO(base64.b64decode(event_chart_image_data.split(',')[1]))
    age_chart_image = BytesIO(base64.b64decode(age_chart_image_data.split(',')[1]))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="event_age_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, leading=22))
    styles.add(ParagraphStyle(name='SubTitle', alignment=1, fontSize=12, leading=15))

    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "img/General_Santos_City_seal.jpg")
    logo = Image(logo_path, width=1*inch, height=1*inch)
    title = Paragraph("Event and Age Distribution Report", styles['CenterTitle'])
    subtitle = Paragraph(f"Date Range: {start_date if start_date != 'ALL' else 'All'} to {end_date if end_date != 'ALL' else 'All'}", styles['SubTitle'])

    title_table = Table([[title], [subtitle]], colWidths=[5.5*inch])
    title_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    header_table = Table([[logo, title_table]], colWidths=[1.5*inch, 5.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Event Attendance Distribution Chart", styles['SubTitle']))
    elements.append(Image(event_chart_image, width=6*inch, height=3*inch))
    elements.append(Spacer(1, 12))

    # Event Data Table
    elements.append(Paragraph("Event Attendance Data", styles['SubTitle']))
    event_data = [
        ['Event Title', 'Date', 'Attendance Count']
    ]
    event_titles = request.POST.getlist('event_titles[]')
    event_dates = request.POST.getlist('event_dates[]')
    event_counts = request.POST.getlist('event_counts[]')

    for title, date, count in zip(event_titles, event_dates, event_counts):
        event_data.append([title, date, count])

    event_table = Table(event_data, colWidths=[3*inch, 2*inch, 2*inch])
    event_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d3d3d3'),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('GRID', (0, 0), (-1, -1), 1, '#000000')
    ]))
    elements.append(event_table)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Freelancer Attendance Age Distribution Chart", styles['SubTitle']))
    elements.append(Image(age_chart_image, width=6*inch, height=3*inch))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Age Distribution Data", styles['SubTitle']))
    age_data = [
        ['Age Range', 'Attendance Count']
    ]
    age_ranges = request.POST.get('age_ranges').split(',')
    age_range_counts = request.POST.get('age_range_counts').split(',')
    for age_range, count in zip(age_ranges, age_range_counts):
        age_data.append([age_range, count])
    age_table = Table(age_data, colWidths=[4*inch, 2*inch])
    age_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d3d3d3'),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('GRID', (0, 0), (-1, -1), 1, '#000000')
    ]))
    elements.append(age_table)
    elements.append(Spacer(1, 12))

    doc.build(elements)

    return response


def render_event_statistics(request, event_id):
    # Fetch the specific event or return a 404 if it doesn't exist
    event = get_object_or_404(Event, id=event_id, is_published=True)
    
    # Retrieve attendance and RSVP details for the event
    attendance_list = Attendance.objects.filter(event=event).values(
        'name', 'date', 'age_range', 'gender', 'pwd', 'four_ps', 'affiliation'
    )
    rsvp_list = RSVP.objects.filter(event=event).values(
        'user__username', 'status', 'timestamp'
    )
    
    # Pass the data to the template
    context = {
        'event_id': event.id,
        'event_title': event.title,
        'event_date': event.start_datetime,
        'event_description': event.description,
        'event_organizer': event.organizer,
        'event_location_type': event.location_type,
        'event_location_name': event.location.name if event.location else 'N/A',
        'event_virtual_platform': event.virtual_details.platform if event.virtual_details else 'N/A',
        'event_qr_code_url': event.qr_code.url if event.qr_code else '',
        'attendance_list': list(attendance_list),
        'rsvp_list': list(rsvp_list),
        'event_image_url': event.image.url if event.image else "",  # Pass the event image URL

    }
    
    return render(request, 'admin/event_statistics.html', context)


@staff_member_required
def export_event_statistics_excel(request, event_id):
    event = get_object_or_404(Event, id=event_id, is_published=True)

    attendances = Attendance.objects.filter(event=event)
    rsvps = RSVP.objects.filter(event=event)
  # Counts
    gender_counts = Counter(attendance.gender for attendance in attendances if attendance.gender)
    status_counts = Counter(rsvp.status for rsvp in rsvps if rsvp.status)
    highest_attendance_day = attendances.values('date').annotate(total=Count('id')).order_by('-total').first()

    wb = Workbook()

    # ➤ Attendance Sheet
    ws1 = wb.active
    ws1.title = "Attendance"
    headers1 = ['Name', 'Date', 'Age Range', 'Gender', 'PWD', '4Ps Member', 'Affiliation']
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for attendance in attendances:
        ws1.append([
            attendance.name,
            attendance.date.strftime('%Y-%m-%d') if attendance.date else '',
            attendance.age_range,
            attendance.gender,
            'Yes' if attendance.pwd else 'No',
            'Yes' if attendance.four_ps else 'No',
            attendance.affiliation
        ])

    # ➤ RSVP Sheet
    ws2 = wb.create_sheet(title="RSVPs")
    headers2 = ['Username', 'Status', 'Timestamp']
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for rsvp in rsvps:
        ws2.append([
            rsvp.user.username if rsvp.user else 'Guest',
            rsvp.status,
            rsvp.timestamp.strftime('%Y-%m-%d %H:%M') if rsvp.timestamp else ''
        ])

    # ➤ Charts Sheet
   # ➤ Charts Sheet
    ws3 = wb.create_sheet(title="Charts")

    def style_header(cell):
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Insights Section
    ws3.merge_cells('A1:B1')
    ws3['A1'] = "Event Insights"
    style_header(ws3['A1'])

    ws3['A3'] = "Total Attendance:"
    ws3['B3'] = attendances.count()

    ws3['A4'] = "Total RSVPs:"
    ws3['B4'] = rsvps.count()

    ws3['A5'] = "Most Common Gender:"
    ws3['B5'] = gender_counts.most_common(1)[0][0] if gender_counts else 'N/A'

    ws3['A6'] = "Most Common RSVP Status:"
    ws3['B6'] = status_counts.most_common(1)[0][0] if status_counts else 'N/A'

    highest_attendance_day = attendances.values('date').annotate(total=Count('id')).order_by('-total').first()
    ws3['A7'] = "Highest Attendance Date:"
    ws3['B7'] = highest_attendance_day['date'].strftime('%Y-%m-%d') if highest_attendance_day else 'N/A'

    ws3.append([])

    # Gender Distribution Table
    gender_table_start_row = ws3.max_row + 2
    ws3.append(['Gender', 'Count'])
    for cell in ws3[gender_table_start_row]:
        style_header(cell)

    for gender, count in gender_counts.items():
        ws3.append([gender, count])

    gender_table_end_row = ws3.max_row

    # ➤ Gender Chart
    gender_chart = BarChart()
    gender_chart.title = "Attendance Gender Distribution"
    gender_chart.y_axis.title = "Count"
    gender_chart.x_axis.title = "Gender"
    gender_chart.style = 13

    gender_data = Reference(ws3, min_col=2, min_row=gender_table_start_row, max_row=gender_table_end_row)
    gender_categories = Reference(ws3, min_col=1, min_row=gender_table_start_row + 1, max_row=gender_table_end_row)
    gender_chart.add_data(gender_data, titles_from_data=True)
    gender_chart.set_categories(gender_categories)
    ws3.add_chart(gender_chart, f"D{gender_table_start_row}")

    ws3.append([])

    # RSVP Status Distribution Table
    rsvp_table_start_row = ws3.max_row + 2
    ws3.append(['RSVP Status', 'Count'])
    for cell in ws3[rsvp_table_start_row]:
        style_header(cell)

    for status, count in status_counts.items():
        ws3.append([status, count])

    rsvp_table_end_row = ws3.max_row

    # ➤ RSVP Status Chart
    rsvp_chart = BarChart()
    rsvp_chart.title = "RSVP Status Distribution"
    rsvp_chart.y_axis.title = "Count"
    rsvp_chart.x_axis.title = "Status"
    rsvp_chart.style = 13

    rsvp_data = Reference(ws3, min_col=2, min_row=rsvp_table_start_row, max_row=rsvp_table_end_row)
    rsvp_categories = Reference(ws3, min_col=1, min_row=rsvp_table_start_row + 1, max_row=rsvp_table_end_row)
    rsvp_chart.add_data(rsvp_data, titles_from_data=True)
    rsvp_chart.set_categories(rsvp_categories)
    ws3.add_chart(rsvp_chart, f"D{rsvp_table_start_row}")


    # ➤ Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'event_statistics_{event.title}_{now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response