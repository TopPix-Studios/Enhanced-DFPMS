from django.contrib import admin
from .models import Event, Attendance, Location, Announcement
from django.core.files.base import ContentFile
# Register your models here.
import csv
from django.http import HttpResponse
from .forms import  EventForm
from django.urls import path
from django.urls import reverse
import qrcode   
from django.shortcuts import render,  redirect, get_object_or_404
from io import BytesIO
from django.utils.html import format_html
from .forms import AnnouncementForm
from profiling.models import Profile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def export_to_csv(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={meta}.csv'
    writer = csv.writer(response)

    writer.writerow(field_names)
    for obj in queryset:
        writer.writerow([getattr(obj, field) for field in field_names])

    return response

export_to_csv.short_description = "Export Selected to CSV"
def export_attendance_sheet(modeladmin, request, queryset):
    if queryset.count() > 1:
        modeladmin.message_user(request, "Please select only one event to export the attendance sheet.")
        return

    event = queryset.first()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Sheet"

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_num = 1

    # Legend Section
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
    ws.cell(row=row_num, column=1, value="Legend:").font = bold_font
    row_num += 1

    ws.cell(row=row_num, column=1, value="Age Range:")
    ws.cell(row=row_num, column=2, value="A = 20 below")
    ws.cell(row=row_num, column=3, value="B = 20-29")
    ws.cell(row=row_num, column=4, value="C = 30-39")
    ws.cell(row=row_num, column=5, value="D = 40-49")
    ws.cell(row=row_num, column=6, value="E = 50+")
    row_num += 1

    ws.cell(row=row_num, column=1, value="Gender:")
    ws.cell(row=row_num, column=2, value="M = Male")
    ws.cell(row=row_num, column=3, value="F = Female")
    ws.cell(row=row_num, column=4, value="O = Other")
    row_num += 2  # Empty row before data

    attendances = Attendance.objects.filter(event=event).order_by('date')
    current_date = None

    for attendance in attendances:
        if current_date != attendance.date:
            current_date = attendance.date
            row_num += 1

            # Date Header
            ws.cell(row=row_num, column=1, value=f"Attendance for {current_date}").font = bold_font
            row_num += 1

            headers = ['Name', 'Logged In', 'Age Range', 'Gender', 'PWD', '4Ps', 'Affiliation', 'Contact', 'Email']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border
            row_num += 1

        # Write attendance row
        values = [
            attendance.name,
            "Yes" if attendance.logged_in else "No",
            attendance.get_age_range_display(),
            attendance.get_gender_display(),
            "Yes" if attendance.pwd else "No",
            "Yes" if attendance.four_ps else "No",
            attendance.affiliation,
            attendance.contact,
            attendance.email
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = center_align
            cell.border = thin_border
        row_num += 1

    # Auto-adjust column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Export as response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{event.title}_attendance.xlsx"'
    wb.save(response)
    return response

export_attendance_sheet.short_description = "Export Attendance Sheet (Per Day)"

class AttendanceInline(admin.TabularInline):
    model = Attendance
    extra = 1  # Number of extra forms to display
    fields = ('user', 'name', 'logged_in')
    readonly_fields = ('name',)
    can_delete = False  # Optional: Prevent deletion of attendance records

def export_rsvp_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "RSVP List"

    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    headers = ['Event Title', 'Name', 'RSVP Status', 'Timestamp', 'Signature']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    row_num = 2  # Start writing data from row 2

    for event in queryset:
        rsvps = event.rsvps.all()
        for rsvp in rsvps:
            try:
                profile = Profile.objects.get(account_id=rsvp.user.account_id)
                name = f"{profile.first_name} {profile.last_name}"
            except Profile.DoesNotExist:
                name = rsvp.user.username

            row = [
                event.title,
                name,
                rsvp.get_status_display(),
                rsvp.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                ""  # Signature (empty for now)
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = center_align
                cell.border = thin_border
            row_num += 1

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=rsvps.xlsx'
    wb.save(response)
    return response

export_rsvp_to_excel.short_description = "Export RSVP to Excel"
def export_attending_rsvp_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attending RSVPs"

    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row
    headers = ['Event Title', 'Name', 'RSVP Status', 'Timestamp', 'Signature']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    row_num = 2

    for event in queryset:
        # Filter only "attending" RSVPs (adjust value if using different status enum/slug)
        rsvps = event.rsvps.filter(status="attending")

        for rsvp in rsvps:
            try:
                profile = Profile.objects.get(account_id=rsvp.user.account_id)
                name = f"{profile.first_name} {profile.last_name}"
            except Profile.DoesNotExist:
                name = rsvp.user.username

            row = [
                event.title,
                name,
                rsvp.get_status_display(),
                rsvp.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                ""  # Signature field
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = center_align
                cell.border = thin_border

            row_num += 1

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Return Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=rsvps_attending.xlsx'
    wb.save(response)
    return response

export_attending_rsvp_to_excel.short_description = "Export RSVP (Attending) to Excel"



class LocationAdmin(admin.ModelAdmin):
    list_display = ['name', 'latitude', 'longitude']
    search_fields = ['name']
    
    class Media:
        css = {
            'all': ('https://unpkg.com/leaflet/dist/leaflet.css',)
        }
        js = (
            'https://unpkg.com/leaflet/dist/leaflet.js',
            'js/location_map.js',  # Replace with the correct path to your JS file
        )

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['latitude'].widget.attrs['id'] = 'latitude-input'
        form.base_fields['longitude'].widget.attrs['id'] = 'longitude-input'
        return form

    def render_change_form(self, request, context, *args, **kwargs):
        context['adminform'].form.fields['name'].widget.attrs.update({'id': 'location-input'})
        return super(LocationAdmin, self).render_change_form(request, context, *args, **kwargs)
    
class EventAdmin(admin.ModelAdmin):
    form = EventForm

    change_list_template = "admin/event_change_list.html"
    list_display = (
        'title',
        'start_datetime',
        'end_datetime',
        'location_type',
        'is_published',
        'is_cancelled',
        'guest_registration_link',
        'view_statistics_link',  # Add the statistics link to list display

    )
    list_filter = (
        'is_published',
        'is_cancelled',
        'start_datetime',
        'tags',
        'location_type'
    )
    search_fields = (
        'title',
        'description',
        'location',
        'tags__name',
        'virtual_platform'
    )
    date_hierarchy = 'start_datetime'
    ordering = ('-start_datetime',)
    actions = [export_to_csv, export_attendance_sheet, export_rsvp_to_excel, export_attending_rsvp_to_excel]
    inlines = [AttendanceInline]

    fieldsets = (
        (None, {
            'fields': ('title', 'description', 'image', 'organizer')
        }),
        ('Event Date', {
            'fields': ('start_datetime', 'end_datetime')
        }),
        ('Location Type', {
            'fields': ('location_type',)
        }),
        ('Physical Location', {
            'fields': ('location', 'latitude', 'longitude'),
            'classes': ('collapse',)  # Collapsed initially
        }),
        ('Virtual Details', {
            'fields': ('virtual_url', 'virtual_platform', 'details'),
            'classes': ('collapse',)  # Collapsed initially
        }),
        ('Status', {
            'fields': ('is_published', 'is_moved', 'is_cancelled','remarks'),
            'classes': ('wide',)
        }),
        ('Tags', {
            'fields': ('tags',)
        }),
    )
    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)

        event_instance = form.instance  # Get the Event this is attached to

        for formset in formsets:
            if formset.model == Attendance:
                for instance in formset.save(commit=False):
                    # Assign event if not already set
                    if not instance.event_id:
                        instance.event = event_instance

                    # Auto-fill name from Profile if not provided
                    if not instance.name and instance.user:
                        try:
                            profile = Profile.objects.get(account_id=instance.user)
                            parts = [profile.first_name]
                            if profile.middle_name:
                                parts.append(profile.middle_name)
                            parts.append(profile.last_name)
                            if profile.suffix:
                                parts.append(profile.suffix)
                            instance.name = " ".join(parts)
                        except Profile.DoesNotExist:
                            instance.name = "No Name"

                    instance.save()
                    
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['event_chart_url'] = '/admin/event-chart/'
        return super(EventAdmin, self).changelist_view(request, extra_context=extra_context)
    
    def view_statistics_link(self, obj):
        """
        Generates a link to the event statistics page.
        """
        url = reverse('render_event_statistics', args=[obj.id])
        return format_html('<a href="{}" target="_blank">View Statistics</a>', url)
    
    view_statistics_link.short_description = "View Statistics"

    def guest_registration_link(self, obj):
        link = reverse('guest_event', args=[obj.id])
        qr_url = reverse('admin:event_qr_code', args=[obj.id])
        return format_html(
            '<a href="{}" target="_blank">Copy Link</a> | <a href="{}" target="_blank">Generate QR Code</a>',
            link,
            qr_url
        )
    guest_registration_link.short_description = 'Guest Link'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'qr-code/<int:event_id>/',
                self.admin_site.admin_view(self.generate_qr_code),
                name='event_qr_code',
            ),
        ]
        return custom_urls + urls

    def generate_qr_code(self, request, event_id):
        event = get_object_or_404(Event, pk=event_id)
        link = request.build_absolute_uri(reverse('guest_event', args=[event.id]))
        qr = qrcode.make(link)
        buffer = BytesIO()
        qr.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename={event.title}_qr.png'
        return response

    def save_model(self, request, obj, form, change):
    
        super().save_model(request, obj, form, change)  # Save the event first

        # Generate and save QR code after the event is created
        if not change:  # Only generate QR code for new events
            link = request.build_absolute_uri(reverse('guest_event', args=[obj.pk]))
            qr = qrcode.make(link)
            buffer = BytesIO()
            qr.save(buffer, format='PNG')
            buffer.seek(0)
            
            # Create a ContentFile from the buffer
            qr_code_content = ContentFile(buffer.getvalue(), name=f"{obj.title}_qr.png")
            obj.qr_code.save(qr_code_content.name, qr_code_content)  # Save the QR code image to the model
            obj.save()  # Save the event again to include the QR code
    class Media:
        css = {
            'all': ('https://unpkg.com/leaflet/dist/leaflet.css',)
        }
        js = (
            'https://unpkg.com/leaflet/dist/leaflet.js',
            'js/map_widget_eve.js',
            'js/event_admin.js',  # Reference to your custom JavaScript file
        )

admin.site.register(Event, EventAdmin)
admin.site.register(Location, LocationAdmin)


class AnnouncementAdmin(admin.ModelAdmin):
    form = AnnouncementForm

    list_display = ('title', 'author', 'created_at', 'address')
    list_filter = ('author', 'tags', 'created_at')
    search_fields = ('title', 'content', 'address', 'author__username', 'tags__name')
    date_hierarchy = 'created_at'
    ordering = ('-created_at',)
    actions = [export_to_csv]
    exclude = ('author',)
    fieldsets = (
        (None, {
            'fields': ('title', 'content', 'image', 'address', 'latitude', 'longitude')
        }),
        ('Tags', {
            'fields': ('tags',)
        }),
    )

    readonly_fields = ('created_at',)

    def save_model(self, request, obj, form, change):
        if not change:  # Only set the author if the object is being created
            obj.author = request.user
        super().save_model(request, obj, form, change)

    class Media:
        css = {
            'all': ('https://unpkg.com/leaflet/dist/leaflet.css',)
        }
        js = (
            'https://unpkg.com/leaflet/dist/leaflet.js',
            'js/map_widget_ann.js',
        )

admin.site.register(Announcement, AnnouncementAdmin)
# admin.site.register(Attendance)